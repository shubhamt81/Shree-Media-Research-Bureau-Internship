import sys
import openpyxl
# sys.append("C:\Users\shubh\AppData\Local\Programs\Python\Python38\Lib\site-packages\openpyxl")

path = "D:\Rough\Python_harish\data_raw.xlsx"
 
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
max_row=sheet_obj.max_row

for i in range(3,max_row+1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    cell_obj.value=cell_obj.value*2
wb_obj.save(path)