from distutils.cmd import Command
from fileinput import filename
from hashlib import new
from tkinter import *
from tkinter import filedialog as fd
from tkinter import messagebox
import sys
from tkinter.messagebox import showinfo
import openpyxl
import os
# sys.append("C:\Users\shubh\AppData\Local\Programs\Python\Python38\Lib\site-packages\openpyxl")
# file operation
# file=""
dict={}

#add caste
def add_caste():
    root=Tk()
    root.title("Add new caste")
    w=Canvas(root, width=400,height=200)
    w.pack()
    canvas_height=200
    canvas_width=400
    T=Text(root,height=1,width=20)
    T.pack()
    T.place(x=100,y=50)
    
    
    def st():
        
        input = str(T.get(1.0,END))
        print(input)
        caste_option.append(input[:-1])
        print(caste_option)
        master.update()
        messagebox.showerror("ALERT",input+" Caste added!")
        root.destroy()
    
    k=Button(root,text="ADD",command=st)
    k.pack()
    k.place(x=100,y=100)
    mainloop()
    


#file opener
def file_opener():
    filetypes = (
        ('All files', '*.*'),
        ('text files', '*.txt')
    )

    filename = fd.askopenfilename(
        title='Open a file',
        initialdir='\\',
        filetypes=filetypes)
    return filename
#############
def save():
    wb_obj.save(path)
    messagebox.showerror("ALERT","FILE SAVED")
#############


def changed():
    msg='Caste of Surname is changed Successfully!\n'+sirname.get()+'->'+caste.get()
    
    if(sirname.get()=="Select Sirname" or caste.get()=="Select Caste"):
        messagebox.showerror("ALERT","SELECT ALL THE VALUES")
        
    else:
        
        r=i
        p=j
        # print(p,r,sirname.get(),caste.get())
        lst=dict[sirname.get()]
        print(lst)
        for k in lst:
            sheet_obj.cell(row=k,column=p).value=caste.get()
        
        messagebox.showinfo("ALERT",msg)
    
    

fil=file_opener()
# print(fil,"f")
f=""
for x in fil:
    if(x=='/'):
        f+='\\'
    else:
        f+=x
path=f
 
#opening the file
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
max_row=sheet_obj.max_row
max_columns=sheet_obj.max_column
print("rows ",max_row,"max_columns ",max_columns)

#select row and columns
j=-1
i=-1
def column_select():
    root=Tk()
    root.title("select columns")
    w=Canvas(root, width=400,height=400)
    w.pack()
    canvas_height=400
    canvas_width=400
    caste_column=[x for x in range(1,max_columns+1)]
    l=StringVar(root)
    l.set("Select Row for caste")
    caste_col=OptionMenu(root,l, *caste_column)
    caste_col.pack()
    caste_col.place(x=250,y=100)

    sirname_column=[x for x in range(1,max_columns+1)]
    m=StringVar(root)
    m.set("Select Row for sirname")
    sirname_col=OptionMenu(root,m, *sirname_column)
    sirname_col.pack()
    sirname_col.place(x=50,y=100)
    def st():
        
        if(l.get()=="Select Row for caste" or m.get()=="Select Row for sirname"):
            messagebox.showwarning("Warning","WARNING COLUMN FOR CASTE AND SIRNAME NOT SELECTED!")
            root.destroy()
        else:
            print("AVAILABLE")
            root.destroy()
            print(l.get(),m.get())
        # return i.get(),j.get()
        
        # mainloop()
    k=Button(root,text="SELECT",command=st)
    k.pack()
    k.place(x=150,y=300)
    mainloop()
    
    return int(m.get()),int(l.get())
    


#calling columns select
i,j=column_select()
print("HERE",i,j)
# OPTIONS FURTHER

#starting canvas

master=Tk()
master.title("Excel quick value set app")
w=Canvas(master, width=400,height=400)
w.pack()
canvas_height=400
canvas_width=400

menubar = Menu(master)
master.config(menu=menubar)

def refresh(self):
    print("refresh")
    self.refresh()
    
file_menu = Menu(
    menubar,
    tearoff=0
)
# file_menu.add_command(label='New')
file_menu.add_command(
    label='Open...'
    ,command=file_opener
)
file_menu.add_command(label='Close')
file_menu.add_separator()
file_menu.add_command(
    label='REFRESH',
    command=refresh
)
file_menu.add_command(
    label='Exit',
    command=master.destroy
)

menubar.add_cascade(
    label="File",
    menu=file_menu
)

add_or_remove=Menu(
    menubar,
    tearoff=0
)
# add_or_remove.add_command(label='Add Sirname')
add_or_remove.add_command(label='Add Caste',command=add_caste)
# add_or_remove.add_command(label='Remove Sirname')
add_or_remove.add_command(label='Remove Caste')

menubar.add_cascade(
    label="ADD/REMOVE",
    menu=add_or_remove
    
)

help_menu = Menu(
    menubar,
    tearoff=0
)

help_menu.add_command(label='Welcome')
help_menu.add_command(
    label='About...'
    # print("CALL SHUBHAM HE MADE THIS")
)
menubar.add_cascade(
    label="Help",
    menu=help_menu
)


#main  window
caste_option=["Bhramin","Banjara","Sonar","Kshatriya","Pathan"]
caste=StringVar(master)
caste.set("Select Caste")
caste_menu=OptionMenu(master,caste, *caste_option)
caste_menu.pack()
caste_menu.place(x=250,y=200)


lst=[]

for k in range(3,max_row):
    cell_obj = sheet_obj.cell(row = k, column =i)
    if(cell_obj.value==None):
                continue
    else:
        lq=list(map(str,cell_obj.value.split()))
        val=lq[-1]
        if(dict.get(val)==None):
            dict[val]=[k]
        else:
            dict[val].append(k)
        #print(cell_obj.value)
        lst.append(str(val))


# print(dict)
sirname_option=list(set(lst))
sirname=StringVar(master)
sirname.set("Select Sirname")
sirname_menu=OptionMenu(master,sirname, *sirname_option)
sirname_menu.pack()
sirname_menu.place(x=50,y=200)

B=Button(text="Change",command=changed)
B.pack()
B.place(x=100,y=300)
T=Button(text="SAVE",command=save)
T.pack()
T.place(x=200,y=300)

mainloop()

